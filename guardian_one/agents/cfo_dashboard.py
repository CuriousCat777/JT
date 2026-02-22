"""CFO Dashboard — Excel Expense Tracker.

Classic Excel expense tracker layout that works for anyone.
Clean tables, Excel formulas, drop-down categories, auto-filters.

Sheets:
    1. Dashboard      — Summary: net worth, income vs spending, accounts, budget, bills
    2. Expenses        — Transaction register with running totals and category dropdowns
    3. Budget          — Budget vs Actual with formulas referencing the Expenses sheet
    4. Bills & Income  — Bills tracker + income log

Usage:
    from guardian_one.agents.cfo_dashboard import generate_dashboard
    generate_dashboard(cfo, "data/dashboard.xlsx")
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

if TYPE_CHECKING:
    from guardian_one.agents.cfo import CFO


# -----------------------------------------------------------------------
# Classic Excel color palette — subtle, professional
# -----------------------------------------------------------------------
_HEADER_BG = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_KPI_BG = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_STRIPE = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_LIGHT_RED = PatternFill(start_color="FCD5D5", end_color="FCD5D5", fill_type="solid")
_LIGHT_GREEN = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
_LIGHT_YELLOW = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")

# Fonts — clean Calibri throughout
_TITLE = Font(name="Calibri", size=16, bold=True, color="1F4E79")
_SUBTITLE = Font(name="Calibri", size=11, color="5B6770")
_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_BODY = Font(name="Calibri", size=11, color="333333")
_BODY_BOLD = Font(name="Calibri", size=11, bold=True, color="333333")
_SECTION = Font(name="Calibri", size=13, bold=True, color="1F4E79")
_KPI_LABEL = Font(name="Calibri", size=10, color="5B6770")
_KPI_VALUE = Font(name="Calibri", size=16, bold=True, color="1F4E79")
_RED_FONT = Font(name="Calibri", size=11, color="C0392B")
_RED_BOLD = Font(name="Calibri", size=11, bold=True, color="C0392B")
_GREEN_FONT = Font(name="Calibri", size=11, color="27AE60")
_GREEN_BOLD = Font(name="Calibri", size=11, bold=True, color="27AE60")

_MONEY = '#,##0.00_);[Red](#,##0.00)'
_PCT = '0.0%'

# Borders — thin gray grid
_GRID = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
_BOTTOM_THICK = Border(
    bottom=Side(style="medium", color="1F4E79"),
)
_TOTAL_BORDER = Border(
    top=Side(style="double", color="1F4E79"),
    bottom=Side(style="medium", color="1F4E79"),
)

# Labels
_ACCOUNT_TYPE_LABELS = {
    "checking": "Checking", "savings": "Savings",
    "credit_card": "Credit Card", "loan": "Loan",
    "investment": "Investments", "retirement": "Retirement",
}

_CATEGORY_LABELS = {
    "income": "Income", "housing": "Housing / Rent",
    "utilities": "Utilities", "food": "Food & Groceries",
    "transport": "Transportation", "medical": "Medical / Health",
    "entertainment": "Shopping & Fun", "education": "Education",
    "insurance": "Insurance", "loan_payment": "Loan Payments",
    "savings": "Savings / Transfers", "charitable": "Donations",
    "other": "Other",
}

# Category list for dropdowns (excludes "income" — that's not an expense)
_EXPENSE_CATEGORIES = [
    "Housing / Rent", "Utilities", "Food & Groceries", "Transportation",
    "Medical / Health", "Shopping & Fun", "Education", "Insurance",
    "Loan Payments", "Savings / Transfers", "Donations", "Other",
]


def generate_dashboard(
    cfo: "CFO",
    output_path: str | Path = "data/dashboard.xlsx",
    password: str | None = None,
    gmail_data: dict[str, Any] | None = None,
) -> Path:
    """Generate the Excel expense tracker dashboard.

    Args:
        cfo: CFO agent instance with financial data.
        output_path: Where to save the .xlsx file.
        password: Lock all sheets with this password (visible but not editable).
        gmail_data: Optional Gmail data (unused in tracker layout, kept for API compat).

    Returns the path to the generated file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _build_dashboard_sheet(wb, cfo)
    _build_expenses_sheet(wb, cfo)
    _build_budget_sheet(wb, cfo)
    _build_bills_income_sheet(wb, cfo)

    if password:
        for ws in wb.worksheets:
            ws.protection.sheet = True
            ws.protection.password = password
            ws.protection.enable()
        wb.security.workbookPassword = password
        wb.security.lockStructure = True

    wb.save(str(output_path))
    return output_path


# -----------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------

def _col_widths(ws: Any, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _header_row(ws: Any, row: int, headers: list[str], start: int = 1) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.font = _HEADER
        c.fill = _HEADER_BG
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _GRID


def _data_cell(ws: Any, row: int, col: int, value: Any, stripe: bool = False) -> Any:
    c = ws.cell(row=row, column=col, value=value)
    c.font = _BODY
    c.border = _GRID
    if stripe:
        c.fill = _STRIPE
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        c.number_format = _MONEY
        c.alignment = Alignment(horizontal="right")
    return c


def _data_row(ws: Any, row: int, values: list[Any], start: int = 1, stripe: bool = False) -> None:
    for i, v in enumerate(values):
        _data_cell(ws, row, start + i, v, stripe)


# -----------------------------------------------------------------------
# Sheet 1: Dashboard
# -----------------------------------------------------------------------

def _build_dashboard_sheet(wb: Workbook, cfo: "CFO") -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "1F4E79"

    _col_widths(ws, {"A": 2, "B": 22, "C": 18, "D": 18, "E": 18, "F": 18, "G": 2})

    now = datetime.now(timezone.utc)
    month_str = now.strftime("%Y-%m")
    month_label = now.strftime("%B %Y")

    # --- Title ---
    ws.merge_cells("B1:F1")
    ws.cell(row=1, column=2, value="Financial Dashboard").font = _TITLE
    ws.cell(row=2, column=2, value=f"Updated: {now.strftime('%b %d, %Y')}").font = _SUBTITLE

    # --- KPI row ---
    net = cfo.net_worth()
    by_type = cfo.balances_by_type()
    income = cfo.income_summary(month_str)
    spending = cfo.spending_summary(month_str)
    total_spent = sum(spending.values())
    left_over = income - total_spent

    kpis = [
        ("Net Worth", net),
        ("Income", income),
        ("Spending", total_spent),
        ("Left Over", left_over),
    ]
    for i, (label, val) in enumerate(kpis):
        col = 2 + i
        lbl = ws.cell(row=4, column=col, value=label)
        lbl.font = _KPI_LABEL
        lbl.fill = _KPI_BG
        lbl.alignment = Alignment(horizontal="center")
        lbl.border = _GRID

        vc = ws.cell(row=5, column=col, value=val)
        vc.font = _KPI_VALUE
        vc.number_format = _MONEY
        vc.fill = _KPI_BG
        vc.alignment = Alignment(horizontal="center")
        vc.border = _GRID

    # Color the "Left Over" KPI
    lo_cell = ws.cell(row=5, column=5)
    if left_over < 0:
        lo_cell.font = Font(name="Calibri", size=16, bold=True, color="C0392B")

    ws.row_dimensions[5].height = 30

    # --- Accounts section ---
    row = 7
    ws.cell(row=row, column=2, value="Accounts").font = _SECTION
    row += 1
    _header_row(ws, row, ["Account", "Type", "Balance", "Bank"], start=2)
    row += 1

    accounts = sorted(cfo._accounts.values(), key=lambda a: -a.balance)
    acct_start = row
    for a in accounts:
        type_label = _ACCOUNT_TYPE_LABELS.get(a.account_type.value, a.account_type.value)
        _data_row(ws, row, [a.name, type_label, a.balance, a.institution],
                  start=2, stripe=(row % 2 == 0))
        # Red for negative balances
        bal = ws.cell(row=row, column=4)
        if a.balance < 0:
            bal.font = _RED_FONT
        row += 1

    # Total with SUM formula
    ws.cell(row=row, column=2, value="Total").font = _BODY_BOLD
    ws.cell(row=row, column=2).border = _TOTAL_BORDER
    if accounts:
        f = ws.cell(row=row, column=4, value=f"=SUM(D{acct_start}:D{row - 1})")
    else:
        f = ws.cell(row=row, column=4, value=net)
    f.number_format = _MONEY
    f.font = _BODY_BOLD
    f.border = _TOTAL_BORDER
    for c in [3, 5]:
        ws.cell(row=row, column=c).border = _TOTAL_BORDER
    row += 2

    # --- Budget Status section ---
    budget_check = cfo.budget_check()
    if budget_check:
        ws.cell(row=row, column=2, value="Budget Status").font = _SECTION
        row += 1
        _header_row(ws, row, ["Category", "Budget", "Spent", "Left", "% Used"], start=2)
        row += 1

        for b in budget_check:
            _data_row(ws, row, [b["label"], b["limit"], b["spent"], b["remaining"], None],
                      start=2, stripe=(row % 2 == 0))
            # % Used cell
            pct = ws.cell(row=row, column=6, value=b["percent_used"] / 100)
            pct.number_format = _PCT
            pct.font = _BODY
            pct.border = _GRID
            pct.alignment = Alignment(horizontal="right")
            if row % 2 == 0:
                pct.fill = _STRIPE

            # Color remaining
            rem = ws.cell(row=row, column=5)
            if b["remaining"] < 0:
                rem.font = _RED_BOLD
            else:
                rem.font = _GREEN_FONT

            # Color status via % Used
            if b["percent_used"] > 100:
                pct.font = _RED_BOLD
            elif b["percent_used"] >= 80:
                pct.font = Font(name="Calibri", size=11, bold=True, color="E67E22")
            row += 1

        # Bar chart — compact
        if len(budget_check) > 1:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Budget vs Actual"
            chart.width = 16
            chart.height = 10

            bdata = Reference(ws, min_col=3, min_row=row - len(budget_check) - 1,
                              max_row=row - 1)
            sdata = Reference(ws, min_col=4, min_row=row - len(budget_check) - 1,
                              max_row=row - 1)
            cats = Reference(ws, min_col=2, min_row=row - len(budget_check),
                             max_row=row - 1)
            chart.add_data(bdata, titles_from_data=True)
            chart.add_data(sdata, titles_from_data=True)
            chart.set_categories(cats)
            chart.series[0].graphicalProperties.solidFill = "85C1E9"
            chart.series[1].graphicalProperties.solidFill = "E74C3C"
            ws.add_chart(chart, f"B{row + 1}")
            row += 17  # Space for chart
        else:
            row += 1

    # --- Upcoming Bills section ---
    overdue = cfo.overdue_bills()
    upcoming = cfo.upcoming_bills(days=14)
    all_upcoming = overdue + upcoming

    if all_upcoming:
        ws.cell(row=row, column=2, value="Upcoming Bills").font = _SECTION
        row += 1
        _header_row(ws, row, ["Bill", "Amount", "Due Date", "Status"], start=2)
        row += 1
        for b in all_upcoming:
            today = now.isoformat()[:10]
            status = "OVERDUE" if b.due_date < today else ("Auto-pay" if b.auto_pay else "Manual")
            _data_row(ws, row, [b.name, b.amount, b.due_date, status],
                      start=2, stripe=(row % 2 == 0))
            if b.due_date < today:
                for c in range(2, 6):
                    ws.cell(row=row, column=c).font = _RED_BOLD
            row += 1

    # Freeze title visible
    ws.freeze_panes = "B4"


# -----------------------------------------------------------------------
# Sheet 2: Expenses (the register)
# -----------------------------------------------------------------------

def _build_expenses_sheet(wb: Workbook, cfo: "CFO") -> None:
    ws = wb.create_sheet("Expenses")
    ws.sheet_properties.tabColor = "E67E22"

    _col_widths(ws, {"A": 12, "B": 40, "C": 14, "D": 22, "E": 22, "F": 16})

    # Header row
    _header_row(ws, 1, ["Date", "Description", "Amount", "Category", "Account", "Running Total"])

    # Sort transactions newest first
    txns = sorted(cfo._transactions, key=lambda t: t.date, reverse=True)

    row = 2
    for tx in txns:
        cat_label = _CATEGORY_LABELS.get(tx.category.value, tx.category.value)
        stripe = row % 2 == 0

        _data_cell(ws, row, 1, tx.date, stripe)
        _data_cell(ws, row, 2, tx.description, stripe)

        # Amount — red for spending, default for income
        amt = _data_cell(ws, row, 3, tx.amount, stripe)
        if tx.amount < 0:
            amt.font = _RED_FONT

        _data_cell(ws, row, 4, cat_label, stripe)
        _data_cell(ws, row, 5, tx.account, stripe)

        # Running total formula (cumulative sum from top)
        rt = ws.cell(row=row, column=6, value=f"=SUM($C$2:C{row})")
        rt.number_format = _MONEY
        rt.font = _BODY
        rt.border = _GRID
        rt.alignment = Alignment(horizontal="right")
        if stripe:
            rt.fill = _STRIPE

        row += 1

    last_data_row = row - 1

    # Total row
    if txns:
        ws.cell(row=row, column=1, value="TOTAL").font = _BODY_BOLD
        ws.cell(row=row, column=1).border = _TOTAL_BORDER
        ws.cell(row=row, column=2).border = _TOTAL_BORDER
        total = ws.cell(row=row, column=3, value=f"=SUM(C2:C{last_data_row})")
        total.number_format = _MONEY
        total.font = _BODY_BOLD
        total.border = _TOTAL_BORDER
        for c in [4, 5, 6]:
            ws.cell(row=row, column=c).border = _TOTAL_BORDER

    # Category dropdown (data validation) on column D
    if txns:
        cat_list = ",".join(_EXPENSE_CATEGORIES)
        dv = DataValidation(
            type="list",
            formula1=f'"{cat_list}"',
            allow_blank=True,
        )
        dv.error = "Pick a category from the list"
        dv.errorTitle = "Invalid Category"
        dv.prompt = "Select a category"
        dv.promptTitle = "Category"
        ws.add_data_validation(dv)
        dv.add(f"D2:D{last_data_row}")

    # Auto-filter
    if txns:
        ws.auto_filter.ref = f"A1:F{last_data_row}"

    # Freeze header row
    ws.freeze_panes = "A2"


# -----------------------------------------------------------------------
# Sheet 3: Budget (formula-driven)
# -----------------------------------------------------------------------

def _build_budget_sheet(wb: Workbook, cfo: "CFO") -> None:
    ws = wb.create_sheet("Budget")
    ws.sheet_properties.tabColor = "8E44AD"

    _col_widths(ws, {"A": 22, "B": 16, "C": 16, "D": 16, "E": 12})

    month_label = datetime.now(timezone.utc).strftime("%B %Y")

    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value=f"Monthly Budget — {month_label}").font = _TITLE

    # Headers
    _header_row(ws, 3, ["Category", "Budget", "Actual", "Remaining", "% Used"])

    budget_check = cfo.budget_check()

    if not budget_check:
        ws.cell(row=5, column=1, value="No budgets set yet.").font = _BODY
        ws.cell(row=6, column=1, value="Use cfo.set_budget('food', 500) to add budgets.").font = _SUBTITLE
        return

    row = 4
    for i, b in enumerate(budget_check):
        stripe = row % 2 == 0
        _data_cell(ws, row, 1, b["label"], stripe)
        _data_cell(ws, row, 2, b["limit"], stripe)

        # Actual — SUMIFS formula referencing Expenses sheet
        cat_label = b["label"]
        actual = ws.cell(row=row, column=3,
                         value=f'=-SUMIFS(Expenses!C:C,Expenses!D:D,"{cat_label}")')
        actual.number_format = _MONEY
        actual.font = _BODY
        actual.border = _GRID
        actual.alignment = Alignment(horizontal="right")
        if stripe:
            actual.fill = _STRIPE

        # Remaining = Budget - Actual
        rem = ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        rem.number_format = _MONEY
        rem.border = _GRID
        rem.alignment = Alignment(horizontal="right")
        if stripe:
            rem.fill = _STRIPE
        # Color: green if positive, red if negative
        # (Can't do conditional font color with static font, so use the computed value)
        if b["remaining"] >= 0:
            rem.font = _GREEN_FONT
        else:
            rem.font = _RED_BOLD

        # % Used = Actual / Budget
        pct = ws.cell(row=row, column=5, value=f"=IF(B{row}>0,C{row}/B{row},0)")
        pct.number_format = _PCT
        pct.font = _BODY
        pct.border = _GRID
        pct.alignment = Alignment(horizontal="right")
        if stripe:
            pct.fill = _STRIPE
        if b["percent_used"] > 100:
            pct.font = _RED_BOLD
        elif b["percent_used"] >= 80:
            pct.font = Font(name="Calibri", size=11, bold=True, color="E67E22")

        row += 1

    # Total row
    data_start = 4
    data_end = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = _BODY_BOLD
    ws.cell(row=row, column=1).border = _TOTAL_BORDER
    for col, letter in [(2, "B"), (3, "C"), (4, "D")]:
        t = ws.cell(row=row, column=col, value=f"=SUM({letter}{data_start}:{letter}{data_end})")
        t.number_format = _MONEY
        t.font = _BODY_BOLD
        t.border = _TOTAL_BORDER
    ws.cell(row=row, column=5).border = _TOTAL_BORDER

    # Bar chart — Budget vs Actual
    if len(budget_check) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Budget vs Actual"
        chart.width = 18
        chart.height = 11

        bdata = Reference(ws, min_col=2, min_row=3, max_row=data_end)
        adata = Reference(ws, min_col=3, min_row=3, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=4, max_row=data_end)

        chart.add_data(bdata, titles_from_data=True)
        chart.add_data(adata, titles_from_data=True)
        chart.set_categories(cats)

        chart.series[0].graphicalProperties.solidFill = "85C1E9"
        chart.series[1].graphicalProperties.solidFill = "E74C3C"

        ws.add_chart(chart, f"A{row + 2}")

    ws.freeze_panes = "A4"


# -----------------------------------------------------------------------
# Sheet 4: Bills & Income
# -----------------------------------------------------------------------

def _build_bills_income_sheet(wb: Workbook, cfo: "CFO") -> None:
    ws = wb.create_sheet("Bills & Income")
    ws.sheet_properties.tabColor = "27AE60"

    _col_widths(ws, {"A": 28, "B": 14, "C": 14, "D": 14, "E": 12, "F": 12})

    now = datetime.now(timezone.utc)
    today = now.isoformat()[:10]
    month_str = now.strftime("%Y-%m")

    # ---- Bills section ----
    ws.cell(row=1, column=1, value="Bills").font = _TITLE
    _header_row(ws, 3, ["Bill", "Amount", "Due Date", "Frequency", "Auto-Pay", "Status"])

    bills = sorted(cfo._bills, key=lambda b: b.due_date)
    row = 4
    for b in bills:
        if b.paid:
            status = "Paid"
        elif b.due_date < today:
            status = "OVERDUE"
        else:
            status = "Upcoming"

        stripe = row % 2 == 0
        _data_row(ws, row, [
            b.name, b.amount, b.due_date,
            b.frequency.title() if b.frequency else "One-time",
            "Yes" if b.auto_pay else "No",
            status,
        ], stripe=stripe)

        if status == "OVERDUE":
            for c in range(1, 7):
                ws.cell(row=row, column=c).font = _RED_BOLD
        elif status == "Paid":
            ws.cell(row=row, column=6).font = _GREEN_FONT
        row += 1

    if not bills:
        ws.cell(row=4, column=1, value="No bills tracked yet.").font = _BODY
        row = 5

    # Monthly total
    bills_end = row - 1
    if bills:
        ws.cell(row=row, column=1, value="Total Monthly Bills").font = _BODY_BOLD
        ws.cell(row=row, column=1).border = _TOTAL_BORDER
        monthly = ws.cell(row=row, column=2,
                          value=f"=SUMPRODUCT((F4:F{bills_end}<>\"Paid\")*(D4:D{bills_end}=\"Monthly\")*B4:B{bills_end})")
        monthly.number_format = _MONEY
        monthly.font = _BODY_BOLD
        monthly.border = _TOTAL_BORDER
        for c in range(3, 7):
            ws.cell(row=row, column=c).border = _TOTAL_BORDER
        row += 1

    # Auto-filter on bills
    if bills:
        ws.auto_filter.ref = f"A3:F{bills_end}"

    row += 2

    # ---- Income section ----
    income_title_row = row
    ws.cell(row=row, column=1, value="Income This Month").font = _TITLE
    row += 1
    _header_row(ws, row, ["Date", "Source", "Amount"], start=1)
    row += 1

    income_txns = [
        tx for tx in cfo._transactions
        if tx.amount > 0 and tx.date.startswith(month_str)
    ]
    income_txns.sort(key=lambda t: t.date, reverse=True)

    income_start = row
    for tx in income_txns:
        stripe = row % 2 == 0
        _data_cell(ws, row, 1, tx.date, stripe)
        _data_cell(ws, row, 2, tx.description, stripe)
        amt = _data_cell(ws, row, 3, tx.amount, stripe)
        amt.font = _GREEN_FONT
        row += 1

    if not income_txns:
        ws.cell(row=row, column=1, value="No income recorded this month.").font = _BODY
        row += 1

    # Income total
    if income_txns:
        income_end = row - 1
        ws.cell(row=row, column=1, value="Total Income").font = _BODY_BOLD
        ws.cell(row=row, column=1).border = _TOTAL_BORDER
        ws.cell(row=row, column=2).border = _TOTAL_BORDER
        t = ws.cell(row=row, column=3, value=f"=SUM(C{income_start}:C{income_end})")
        t.number_format = _MONEY
        t.font = _GREEN_BOLD
        t.border = _TOTAL_BORDER

    ws.freeze_panes = "A4"
