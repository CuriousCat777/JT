"""CFO — Financial Management Agent.

Responsibilities:
- Sync with Rocket Money (account unification via API or CSV)
- Dashboard: income, expenses, loans, savings
- Bill alerts and payment confirmations
- Tax optimisation (retirement, charitable giving)
- Scenario planning (home purchase, retirement)
- Encrypted financial records with offline backup support
"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from pathlib import Path
from typing import Any

from guardian_one.core.audit import AuditLog, Severity
from guardian_one.core.base_agent import AgentReport, AgentStatus, BaseAgent
from guardian_one.core.config import AgentConfig
from guardian_one.integrations.financial_sync import (
    EmpowerProvider,
    PlaidProvider,
    RocketMoneyProvider,
    SyncedAccount,
    SyncedTransaction,
)


class AccountType(Enum):
    CHECKING = "checking"
    SAVINGS = "savings"
    CREDIT_CARD = "credit_card"
    LOAN = "loan"
    INVESTMENT = "investment"
    RETIREMENT = "retirement"


class TransactionCategory(Enum):
    INCOME = "income"
    HOUSING = "housing"
    UTILITIES = "utilities"
    FOOD = "food"
    TRANSPORT = "transport"
    MEDICAL = "medical"
    ENTERTAINMENT = "entertainment"
    EDUCATION = "education"
    INSURANCE = "insurance"
    LOAN_PAYMENT = "loan_payment"
    SAVINGS = "savings"
    CHARITABLE = "charitable"
    OTHER = "other"


@dataclass
class Account:
    name: str
    account_type: AccountType
    balance: float
    institution: str = ""
    last_synced: str = field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )


@dataclass
class Transaction:
    date: str
    description: str
    amount: float  # Positive = inflow, negative = outflow
    category: TransactionCategory = TransactionCategory.OTHER
    account: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class Bill:
    name: str
    amount: float
    due_date: str  # ISO date
    recurring: bool = True
    frequency: str = "monthly"
    auto_pay: bool = False
    paid: bool = False


@dataclass
class Budget:
    """Monthly spending limit for a category."""
    category: str         # TransactionCategory value
    limit: float          # Max spend per month
    label: str = ""       # Friendly name (auto-set if blank)


@dataclass
class NetWorthSnapshot:
    """Point-in-time net worth record for trend tracking."""
    date: str             # ISO date
    net_worth: float
    by_type: dict[str, float] = field(default_factory=dict)


@dataclass
class Scenario:
    """A financial scenario for planning purposes."""
    name: str
    description: str
    assumptions: dict[str, Any] = field(default_factory=dict)
    projections: dict[str, Any] = field(default_factory=dict)


class CFO(BaseAgent):
    """Financial management agent for Jeremy."""

    def __init__(
        self,
        config: AgentConfig,
        audit: AuditLog,
        data_dir: Path | str = "data",
        rocket_money: RocketMoneyProvider | None = None,
        plaid: PlaidProvider | None = None,
    ) -> None:
        super().__init__(config, audit)
        self._accounts: dict[str, Account] = {}
        self._transactions: list[Transaction] = []
        self._bills: list[Bill] = []
        self._budgets: list[Budget] = []
        self._net_worth_history: list[NetWorthSnapshot] = []
        self._scenarios: dict[str, Scenario] = {}
        self._data_dir = Path(data_dir)
        self._ledger_path = self._data_dir / "cfo_ledger.json"
        self._rocket_money = rocket_money or RocketMoneyProvider()
        self._rm_connected = False
        self._empower = EmpowerProvider()
        self._empower_connected = False
        self._plaid = plaid or PlaidProvider(
            token_store_path=self._data_dir / "plaid_tokens.json",
        )
        self._plaid_connected = False
        self._last_sync: str = ""

    def initialize(self) -> None:
        self._set_status(AgentStatus.IDLE)
        loaded = self._load_ledger()

        # Attempt Rocket Money connection
        if self._rocket_money.has_credentials:
            self._rm_connected = self._rocket_money.authenticate()
        else:
            self._rm_connected = False

        # Attempt Empower connection
        if self._empower.has_credentials:
            self._empower_connected = self._empower.authenticate()
        else:
            self._empower_connected = False

        # Attempt Plaid connection
        if self._plaid.has_credentials:
            self._plaid_connected = self._plaid.authenticate()
        else:
            self._plaid_connected = False

        self.log("initialized", details={
            "accounts": len(self._accounts),
            "transactions": len(self._transactions),
            "bills": len(self._bills),
            "loaded_from_disk": loaded,
            "rocket_money_connected": self._rm_connected,
            "rocket_money_has_key": self._rocket_money.has_credentials,
            "empower_connected": self._empower_connected,
            "empower_has_key": self._empower.has_credentials,
            "plaid_connected": self._plaid_connected,
            "plaid_institutions": len(self._plaid.connected_institutions),
        })

    # ------------------------------------------------------------------
    # Persistence — save & load financial state
    # ------------------------------------------------------------------

    def _load_ledger(self) -> bool:
        """Load financial data from the ledger file on disk.

        Returns True if data was loaded, False otherwise.
        The ledger file is a JSON file stored in the data directory.
        """
        if not self._ledger_path.exists():
            return False
        try:
            raw = json.loads(self._ledger_path.read_text())
            self._load_accounts(raw.get("accounts", []))
            self._load_transactions(raw.get("transactions", []))
            self._load_bills(raw.get("bills", []))
            self._load_budgets(raw.get("budgets", []))
            self._load_net_worth_history(raw.get("net_worth_history", []))
            self.log("ledger_loaded", details={
                "path": str(self._ledger_path),
                "accounts": len(self._accounts),
                "transactions": len(self._transactions),
                "bills": len(self._bills),
                "budgets": len(self._budgets),
                "net_worth_snapshots": len(self._net_worth_history),
            })
            return True
        except (json.JSONDecodeError, KeyError, TypeError) as exc:
            self.log(
                "ledger_load_error",
                severity=Severity.ERROR,
                details={"error": str(exc)},
            )
            return False

    def _load_accounts(self, entries: list[dict[str, Any]]) -> None:
        for entry in entries:
            acct = Account(
                name=entry["name"],
                account_type=AccountType(entry["account_type"]),
                balance=entry["balance"],
                institution=entry.get("institution", ""),
                last_synced=entry.get("last_synced", datetime.now(timezone.utc).isoformat()),
            )
            self._accounts[acct.name] = acct

    def _load_transactions(self, entries: list[dict[str, Any]]) -> None:
        for entry in entries:
            tx = Transaction(
                date=entry["date"],
                description=entry["description"],
                amount=entry["amount"],
                category=TransactionCategory(entry.get("category", "other")),
                account=entry.get("account", ""),
                metadata=entry.get("metadata", {}),
            )
            self._transactions.append(tx)

    def _load_bills(self, entries: list[dict[str, Any]]) -> None:
        for entry in entries:
            bill = Bill(
                name=entry["name"],
                amount=entry["amount"],
                due_date=entry["due_date"],
                recurring=entry.get("recurring", True),
                frequency=entry.get("frequency", "monthly"),
                auto_pay=entry.get("auto_pay", False),
                paid=entry.get("paid", False),
            )
            self._bills.append(bill)

    def _load_budgets(self, entries: list[dict[str, Any]]) -> None:
        for entry in entries:
            self._budgets.append(Budget(
                category=entry["category"],
                limit=entry["limit"],
                label=entry.get("label", ""),
            ))

    def _load_net_worth_history(self, entries: list[dict[str, Any]]) -> None:
        for entry in entries:
            self._net_worth_history.append(NetWorthSnapshot(
                date=entry["date"],
                net_worth=entry["net_worth"],
                by_type=entry.get("by_type", {}),
            ))

    def save_ledger(self) -> None:
        """Persist current financial state to disk."""
        data = {
            "saved_at": datetime.now(timezone.utc).isoformat(),
            "accounts": [
                {
                    "name": a.name,
                    "account_type": a.account_type.value,
                    "balance": a.balance,
                    "institution": a.institution,
                    "last_synced": a.last_synced,
                }
                for a in self._accounts.values()
            ],
            "transactions": [
                {
                    "date": tx.date,
                    "description": tx.description,
                    "amount": tx.amount,
                    "category": tx.category.value,
                    "account": tx.account,
                    "metadata": tx.metadata,
                }
                for tx in self._transactions
            ],
            "bills": [
                {
                    "name": b.name,
                    "amount": b.amount,
                    "due_date": b.due_date,
                    "recurring": b.recurring,
                    "frequency": b.frequency,
                    "auto_pay": b.auto_pay,
                    "paid": b.paid,
                }
                for b in self._bills
            ],
            "budgets": [
                {
                    "category": b.category,
                    "limit": b.limit,
                    "label": b.label,
                }
                for b in self._budgets
            ],
            "net_worth_history": [
                {
                    "date": s.date,
                    "net_worth": s.net_worth,
                    "by_type": s.by_type,
                }
                for s in self._net_worth_history
            ],
        }
        self._data_dir.mkdir(parents=True, exist_ok=True)
        self._ledger_path.write_text(json.dumps(data, indent=2))
        self.log("ledger_saved", details={
            "path": str(self._ledger_path),
            "accounts": len(self._accounts),
            "transactions": len(self._transactions),
            "bills": len(self._bills),
        })

    def add_account(self, account: Account, persist: bool = True) -> None:
        self._accounts[account.name] = account
        self.log("account_added", details={"name": account.name, "type": account.account_type.value})
        if persist:
            self.save_ledger()

    def clean_ledger(
        self,
        strip_sandbox: bool = True,
        strip_rm_goals: bool = True,
        strip_zero_dupes: bool = True,
        dry_run: bool = False,
    ) -> dict[str, Any]:
        """Remove sandbox data, Rocket Money goal buckets, and zero-balance duplicates.

        Returns a report of what was (or would be) removed.
        """
        removed_accounts: list[str] = []
        removed_transactions: list[str] = []

        # Known sandbox institutions (Plaid test data)
        sandbox_institutions = {"First Platypus Bank", "Platypus Bank"}

        # Known Rocket Money goal bucket names
        rm_goal_keywords = {"fund", "pay off", "save for"}

        accounts_before = len(self._accounts)
        txns_before = len(self._transactions)

        # --- Strip sandbox accounts ---
        if strip_sandbox:
            sandbox_names = [
                name for name, acct in self._accounts.items()
                if acct.institution in sandbox_institutions
            ]
            for name in sandbox_names:
                removed_accounts.append(f"[sandbox] {name} ({self._accounts[name].institution})")
                if not dry_run:
                    del self._accounts[name]

            # Strip transactions referencing sandbox accounts
            sandbox_account_names = {n.split(" (")[0] for n in sandbox_names}  # match partial
            sandbox_txns = [
                tx for tx in self._transactions
                if any(s in tx.account for s in sandbox_institutions)
                or any(s in tx.account for s in sandbox_account_names)
                or any(s in tx.description for s in sandbox_institutions)
            ]
            for tx in sandbox_txns:
                removed_transactions.append(f"[sandbox] {tx.date} {tx.description} ${tx.amount:.2f}")
            if not dry_run:
                self._transactions = [tx for tx in self._transactions if tx not in sandbox_txns]

        # --- Strip Rocket Money goal buckets (zero-balance pseudo-accounts) ---
        if strip_rm_goals:
            rm_names = [
                name for name, acct in self._accounts.items()
                if acct.institution == "Rocket Money"
                and acct.balance == 0
                and any(kw in name.lower() for kw in rm_goal_keywords)
            ]
            for name in rm_names:
                removed_accounts.append(f"[rm_goal] {name}")
                if not dry_run:
                    del self._accounts[name]

        # --- Strip zero-balance duplicates ---
        # If two accounts share the same institution and one has a balance
        # while the other is $0 with a similar name, remove the zero one.
        if strip_zero_dupes:
            by_inst: dict[str, list[str]] = {}
            for name, acct in self._accounts.items():
                by_inst.setdefault(acct.institution, []).append(name)

            for inst, names in by_inst.items():
                if len(names) < 2:
                    continue
                has_balance = [n for n in names if self._accounts[n].balance != 0]
                zeroes = [n for n in names if self._accounts[n].balance == 0]
                for z in zeroes:
                    # If there's a non-zero account at the same institution, the zero
                    # is likely a stale duplicate
                    if has_balance and z not in [r.split("] ", 1)[-1] for r in removed_accounts]:
                        removed_accounts.append(f"[zero_dupe] {z} ({inst})")
                        if not dry_run:
                            del self._accounts[z]

        if not dry_run:
            self.save_ledger()
            self.log("ledger_cleaned", details={
                "accounts_removed": len(removed_accounts),
                "transactions_removed": len(removed_transactions),
                "accounts_before": accounts_before,
                "accounts_after": len(self._accounts),
                "transactions_before": txns_before,
                "transactions_after": len(self._transactions),
            })

        return {
            "dry_run": dry_run,
            "accounts_removed": removed_accounts,
            "transactions_removed": removed_transactions,
            "accounts_before": accounts_before,
            "accounts_after": accounts_before - len(removed_accounts) if dry_run else len(self._accounts),
            "transactions_before": txns_before,
            "transactions_after": txns_before - len(removed_transactions) if dry_run else len(self._transactions),
        }

    # ------------------------------------------------------------------
    # Account management
    # ------------------------------------------------------------------

    def get_account(self, name: str) -> Account | None:
        return self._accounts.get(name)

    def net_worth(self) -> float:
        return round(sum(a.balance for a in self._accounts.values()), 2)

    def balances_by_type(self) -> dict[str, float]:
        totals: dict[str, float] = {}
        for account in self._accounts.values():
            key = account.account_type.value
            totals[key] = totals.get(key, 0) + account.balance
        return {k: round(v, 2) for k, v in totals.items()}

    # ------------------------------------------------------------------
    # Transaction tracking
    # ------------------------------------------------------------------

    def record_transaction(self, tx: Transaction, persist: bool = True) -> None:
        self._transactions.append(tx)
        if persist:
            self.save_ledger()

    def spending_summary(self, month: str | None = None) -> dict[str, float]:
        """Summarise spending by category.  month format: 'YYYY-MM'."""
        totals: dict[str, float] = {}
        for tx in self._transactions:
            if month and not tx.date.startswith(month):
                continue
            if tx.amount < 0:  # outflow
                key = tx.category.value
                totals[key] = totals.get(key, 0) + abs(tx.amount)
        return totals

    def income_summary(self, month: str | None = None) -> float:
        return sum(
            tx.amount for tx in self._transactions
            if tx.amount > 0 and (month is None or tx.date.startswith(month))
        )

    # ------------------------------------------------------------------
    # Bill management
    # ------------------------------------------------------------------

    def add_bill(self, bill: Bill, persist: bool = True) -> None:
        self._bills.append(bill)
        self.log("bill_added", details={"name": bill.name, "due": bill.due_date})
        if persist:
            self.save_ledger()

    def upcoming_bills(self, days: int = 7) -> list[Bill]:
        now = datetime.now(timezone.utc)
        cutoff = now.isoformat()[:10]
        end = (now + timedelta(days=days)).isoformat()[:10]
        return [
            b for b in self._bills
            if not b.paid and cutoff <= b.due_date <= end
        ]

    def overdue_bills(self) -> list[Bill]:
        today = datetime.now(timezone.utc).isoformat()[:10]
        return [b for b in self._bills if not b.paid and b.due_date < today]

    # ------------------------------------------------------------------
    # Budget tracking
    # ------------------------------------------------------------------

    _CATEGORY_FRIENDLY = {
        "income": "Income",
        "housing": "Housing / Rent",
        "utilities": "Utilities",
        "food": "Food & Groceries",
        "transport": "Transportation",
        "medical": "Medical / Health",
        "entertainment": "Shopping & Fun",
        "education": "Education",
        "insurance": "Insurance",
        "loan_payment": "Loan Payments",
        "savings": "Savings / Transfers",
        "charitable": "Donations",
        "other": "Other",
    }

    def set_budget(self, category: str, limit: float, label: str = "", persist: bool = True) -> Budget:
        """Set a monthly spending limit for a category.

        If a budget already exists for this category, it gets updated.
        """
        if not label:
            label = self._CATEGORY_FRIENDLY.get(category, category.replace("_", " ").title())

        # Update existing or create new
        for b in self._budgets:
            if b.category == category:
                b.limit = limit
                b.label = label
                if persist:
                    self.save_ledger()
                self.log("budget_updated", details={"category": category, "limit": limit})
                return b

        budget = Budget(category=category, limit=limit, label=label)
        self._budgets.append(budget)
        if persist:
            self.save_ledger()
        self.log("budget_set", details={"category": category, "limit": limit})
        return budget

    def remove_budget(self, category: str, persist: bool = True) -> bool:
        """Remove a budget for a category."""
        before = len(self._budgets)
        self._budgets = [b for b in self._budgets if b.category != category]
        if persist and len(self._budgets) != before:
            self.save_ledger()
        return len(self._budgets) != before

    def budget_check(self, month: str | None = None) -> list[dict[str, Any]]:
        """Check spending against budgets for a given month.

        Returns a list of results — one per budget — showing:
        - category, limit, spent, remaining, over_budget, percent_used
        Plain language: "You spent $X of your $Y food budget (Z%)"
        """
        if month is None:
            month = datetime.now(timezone.utc).strftime("%Y-%m")

        spending = self.spending_summary(month)
        results: list[dict[str, Any]] = []

        for b in self._budgets:
            spent = spending.get(b.category, 0)
            remaining = b.limit - spent
            pct = (spent / b.limit * 100) if b.limit > 0 else 0

            if spent > b.limit:
                status = "over"
            elif pct >= 80:
                status = "warning"
            else:
                status = "ok"

            results.append({
                "category": b.category,
                "label": b.label or self._CATEGORY_FRIENDLY.get(b.category, b.category),
                "limit": b.limit,
                "spent": round(spent, 2),
                "remaining": round(remaining, 2),
                "percent_used": round(pct, 1),
                "over_budget": spent > b.limit,
                "status": status,
            })

        return sorted(results, key=lambda r: -r["percent_used"])

    def budget_alerts(self, month: str | None = None) -> list[str]:
        """Get plain-English budget alerts for any over/near-limit categories."""
        alerts: list[str] = []
        for r in self.budget_check(month):
            label = r["label"]
            if r["status"] == "over":
                over_by = abs(r["remaining"])
                alerts.append(
                    f"OVER BUDGET: {label} — spent ${r['spent']:,.2f} "
                    f"of ${r['limit']:,.2f} limit (${over_by:,.2f} over)"
                )
            elif r["status"] == "warning":
                alerts.append(
                    f"Heads up: {label} — ${r['spent']:,.2f} of "
                    f"${r['limit']:,.2f} ({r['percent_used']:.0f}% used)"
                )
        return alerts

    # ------------------------------------------------------------------
    # Net worth history
    # ------------------------------------------------------------------

    def record_net_worth(self, persist: bool = True) -> NetWorthSnapshot:
        """Take a snapshot of current net worth and save it to history.

        Only records one snapshot per day (skips if today already recorded).
        """
        today = datetime.now(timezone.utc).isoformat()[:10]

        # Skip if already recorded today
        for s in self._net_worth_history:
            if s.date == today:
                s.net_worth = self.net_worth()
                s.by_type = self.balances_by_type()
                if persist:
                    self.save_ledger()
                return s

        snapshot = NetWorthSnapshot(
            date=today,
            net_worth=self.net_worth(),
            by_type=self.balances_by_type(),
        )
        self._net_worth_history.append(snapshot)
        if persist:
            self.save_ledger()
        self.log("net_worth_recorded", details={
            "date": today, "net_worth": snapshot.net_worth,
        })
        return snapshot

    def net_worth_trend(self, months: int = 12) -> list[dict[str, Any]]:
        """Get net worth history for the last N months.

        Returns a list of snapshots sorted by date.
        """
        if not self._net_worth_history:
            return []

        cutoff = (datetime.now(timezone.utc) - timedelta(days=months * 30)).isoformat()[:10]
        return [
            {"date": s.date, "net_worth": s.net_worth, "by_type": s.by_type}
            for s in sorted(self._net_worth_history, key=lambda s: s.date)
            if s.date >= cutoff
        ]

    # ------------------------------------------------------------------
    # Transaction verification (cross-check bank data)
    # ------------------------------------------------------------------

    def verify_transactions(self, days: int = 7) -> dict[str, Any]:
        """Independently verify recent transactions for issues.

        Checks for:
        - Duplicate charges (same amount + merchant on the same day)
        - Unusually large transactions (> 3x your average for that category)
        - Unknown merchants you've never seen before
        - Round-number charges that could be unauthorized (e.g. $100.00 exactly)

        Returns a plain-English report with flagged items.
        """
        cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()[:10]
        recent = [tx for tx in self._transactions if tx.date >= cutoff]

        if not recent:
            return {
                "checked": 0,
                "issues": [],
                "summary": "No recent transactions to verify.",
                "status": "ok",
            }

        issues: list[dict[str, Any]] = []

        # 1. Duplicate detection (same amount + similar description + same day)
        by_day: dict[str, list] = {}
        for tx in recent:
            key = tx.date
            by_day.setdefault(key, []).append(tx)

        for day, txns in by_day.items():
            seen: dict[str, list] = {}
            for tx in txns:
                # Group by rounded amount + first word of description
                sig = f"{abs(tx.amount):.2f}|{tx.description.split()[0].lower() if tx.description else ''}"
                seen.setdefault(sig, []).append(tx)
            for sig, group in seen.items():
                if len(group) > 1:
                    issues.append({
                        "type": "duplicate",
                        "severity": "warning",
                        "description": (
                            f"Possible duplicate: {group[0].description} "
                            f"(${abs(group[0].amount):.2f}) appears {len(group)}x on {day}"
                        ),
                        "transactions": [
                            {"date": t.date, "desc": t.description, "amount": t.amount}
                            for t in group
                        ],
                    })

        # 2. Unusually large transactions
        # Build category averages from all history
        cat_amounts: dict[str, list[float]] = {}
        for tx in self._transactions:
            if tx.amount < 0:  # Only outflows
                cat_amounts.setdefault(tx.category.value, []).append(abs(tx.amount))

        for tx in recent:
            if tx.amount >= 0:
                continue
            cat = tx.category.value
            amounts = cat_amounts.get(cat, [])
            if len(amounts) >= 5:  # Need enough data
                avg = sum(amounts) / len(amounts)
                if abs(tx.amount) > avg * 3 and abs(tx.amount) > 50:
                    cat_label = self._CATEGORY_FRIENDLY.get(cat, cat)
                    issues.append({
                        "type": "unusual_amount",
                        "severity": "warning",
                        "description": (
                            f"Unusually large: {tx.description} — ${abs(tx.amount):,.2f} "
                            f"(your average {cat_label} is ${avg:,.2f})"
                        ),
                        "transaction": {"date": tx.date, "desc": tx.description, "amount": tx.amount},
                    })

        # 3. Unknown merchants (first-time senders)
        all_merchants = set()
        for tx in self._transactions:
            if tx.date < cutoff and tx.description:
                all_merchants.add(tx.description.lower().strip())

        for tx in recent:
            if tx.amount < 0 and tx.description:
                if tx.description.lower().strip() not in all_merchants and abs(tx.amount) > 20:
                    issues.append({
                        "type": "new_merchant",
                        "severity": "info",
                        "description": (
                            f"First time: {tx.description} — ${abs(tx.amount):,.2f} on {tx.date}"
                        ),
                        "transaction": {"date": tx.date, "desc": tx.description, "amount": tx.amount},
                    })

        # 4. Suspiciously round numbers
        for tx in recent:
            if tx.amount < 0 and abs(tx.amount) >= 100:
                if abs(tx.amount) == round(abs(tx.amount)):
                    # Exact round number
                    issues.append({
                        "type": "round_amount",
                        "severity": "info",
                        "description": (
                            f"Round number: {tx.description} — exactly ${abs(tx.amount):,.0f} on {tx.date}"
                        ),
                        "transaction": {"date": tx.date, "desc": tx.description, "amount": tx.amount},
                    })

        warnings = [i for i in issues if i["severity"] == "warning"]
        infos = [i for i in issues if i["severity"] == "info"]

        if warnings:
            status = "needs_attention"
            summary = f"{len(warnings)} issue(s) need your attention, {len(infos)} info item(s)."
        elif infos:
            status = "review"
            summary = f"No issues found. {len(infos)} item(s) for your review."
        else:
            status = "ok"
            summary = f"All {len(recent)} recent transactions look normal."

        return {
            "checked": len(recent),
            "days": days,
            "issues": issues,
            "warnings": len(warnings),
            "info_items": len(infos),
            "summary": summary,
            "status": status,
            "verified_at": datetime.now(timezone.utc).isoformat(),
        }

    def verify_bills_paid(self) -> list[dict[str, Any]]:
        """Cross-check bills against transactions to confirm payments.

        For each bill, looks for a matching transaction (close amount,
        around the due date).  Returns status per bill.
        """
        results: list[dict[str, Any]] = []

        for bill in self._bills:
            if bill.paid:
                results.append({
                    "bill": bill.name,
                    "amount": bill.amount,
                    "due": bill.due_date,
                    "status": "confirmed_paid",
                    "message": f"{bill.name} — marked as paid.",
                })
                continue

            # Look for a matching transaction near the due date
            due_dt = bill.due_date
            found = False
            for tx in self._transactions:
                # Match: within 5 days of due date, similar amount (within 10%)
                if not tx.date:
                    continue
                day_diff = abs(
                    (datetime.fromisoformat(tx.date) - datetime.fromisoformat(due_dt)).days
                ) if tx.date >= "2000" and due_dt >= "2000" else 999

                if day_diff <= 5 and abs(tx.amount) > 0:
                    amount_diff = abs(abs(tx.amount) - bill.amount) / bill.amount if bill.amount > 0 else 999
                    if amount_diff < 0.10:  # Within 10%
                        found = True
                        results.append({
                            "bill": bill.name,
                            "amount": bill.amount,
                            "due": bill.due_date,
                            "status": "likely_paid",
                            "matched_transaction": {
                                "date": tx.date,
                                "description": tx.description,
                                "amount": tx.amount,
                            },
                            "message": (
                                f"{bill.name} — likely paid via "
                                f"{tx.description} (${abs(tx.amount):,.2f} on {tx.date})"
                            ),
                        })
                        break

            if not found:
                today = datetime.now(timezone.utc).isoformat()[:10]
                if bill.due_date < today:
                    status = "overdue_unverified"
                    msg = f"{bill.name} — OVERDUE (${bill.amount:,.2f} due {bill.due_date}), no matching payment found!"
                else:
                    status = "pending"
                    msg = f"{bill.name} — due {bill.due_date} (${bill.amount:,.2f}), not yet paid."

                results.append({
                    "bill": bill.name,
                    "amount": bill.amount,
                    "due": bill.due_date,
                    "status": status,
                    "message": msg,
                })

        return results

    def daily_review(self, gmail_data: dict[str, Any] | None = None) -> dict[str, Any]:
        """Run the full CFO daily check.

        Combines:
        1. Transaction verification (cross-check bank data)
        2. Bill payment verification
        3. Budget check
        4. Financial email summary (if gmail_data provided)

        Returns a complete daily review report suitable for
        the Excel dashboard "Daily Check" sheet.
        """
        now = datetime.now(timezone.utc)
        review: dict[str, Any] = {
            "date": now.isoformat()[:10],
            "generated_at": now.isoformat(),
        }

        # 1. Transaction verification
        tx_verify = self.verify_transactions(days=7)
        review["transactions"] = tx_verify

        # 2. Bill payment verification
        bill_verify = self.verify_bills_paid()
        review["bills"] = {
            "results": bill_verify,
            "paid": len([b for b in bill_verify if b["status"] in ("confirmed_paid", "likely_paid")]),
            "pending": len([b for b in bill_verify if b["status"] == "pending"]),
            "overdue": len([b for b in bill_verify if b["status"] == "overdue_unverified"]),
        }

        # 3. Budget check
        budget = self.budget_check()
        budget_over = [b for b in budget if b["status"] == "over"]
        budget_warn = [b for b in budget if b["status"] == "warning"]
        review["budget"] = {
            "results": budget,
            "over_budget": len(budget_over),
            "warnings": len(budget_warn),
            "on_track": len(budget) - len(budget_over) - len(budget_warn),
        }

        # 4. Financial emails (if Gmail data provided)
        if gmail_data:
            review["emails"] = gmail_data
        else:
            review["emails"] = {"available": False, "note": "Run --gmail to include email review"}

        # Overall status
        issues = 0
        if tx_verify["status"] == "needs_attention":
            issues += tx_verify["warnings"]
        issues += review["bills"]["overdue"]
        issues += len(budget_over)

        if issues > 0:
            review["overall_status"] = "needs_attention"
            review["overall_message"] = f"{issues} item(s) need your attention today."
        elif tx_verify["status"] == "review" or budget_warn:
            review["overall_status"] = "review"
            review["overall_message"] = "No urgent issues — a few items to review when you have time."
        else:
            review["overall_status"] = "all_clear"
            review["overall_message"] = "Everything looks good today."

        self.log("daily_review", details={
            "status": review["overall_status"],
            "tx_issues": tx_verify.get("warnings", 0),
            "bill_overdue": review["bills"]["overdue"],
            "budget_over": len(budget_over),
        })
        return review

    # ------------------------------------------------------------------
    # Tax optimisation
    # ------------------------------------------------------------------

    def tax_recommendations(self) -> list[str]:
        """Generate tax optimisation recommendations."""
        recs: list[str] = []

        # Retirement contribution check
        retirement = [a for a in self._accounts.values() if a.account_type == AccountType.RETIREMENT]
        if not retirement:
            recs.append("No retirement accounts detected. Consider opening a 401(k) or IRA to reduce taxable income.")
        else:
            total = sum(a.balance for a in retirement)
            recs.append(f"Retirement balance: ${total:,.2f}. Ensure you are maximising annual contribution limits.")

        # Charitable giving
        charitable_spend = sum(
            abs(tx.amount) for tx in self._transactions
            if tx.category == TransactionCategory.CHARITABLE
        )
        if charitable_spend > 0:
            recs.append(f"Charitable giving: ${charitable_spend:,.2f} — ensure receipts are filed for deduction.")
        else:
            recs.append("Consider charitable donations for potential tax deductions.")

        # Student loan interest
        loan_payments = sum(
            abs(tx.amount) for tx in self._transactions
            if tx.category == TransactionCategory.LOAN_PAYMENT
        )
        if loan_payments > 0:
            recs.append(f"Loan payments: ${loan_payments:,.2f}. Student loan interest may be deductible (up to $2,500/year).")

        return recs

    # ------------------------------------------------------------------
    # Scenario planning
    # ------------------------------------------------------------------

    def create_scenario(self, scenario: Scenario) -> None:
        self._scenarios[scenario.name] = scenario
        self.log("scenario_created", details={"name": scenario.name})

    def home_purchase_scenario(
        self,
        target_price: float,
        down_payment_pct: float = 0.20,
        interest_rate: float = 0.065,
        term_years: int = 30,
    ) -> dict[str, Any]:
        """Simple home purchase affordability projection."""
        down = target_price * down_payment_pct
        loan = target_price - down
        monthly_rate = interest_rate / 12
        n_payments = term_years * 12

        if monthly_rate > 0:
            monthly_payment = loan * (monthly_rate * (1 + monthly_rate) ** n_payments) / (
                (1 + monthly_rate) ** n_payments - 1
            )
        else:
            monthly_payment = loan / n_payments

        current_savings = sum(
            a.balance for a in self._accounts.values()
            if a.account_type in (AccountType.SAVINGS, AccountType.CHECKING)
        )

        result = {
            "target_price": target_price,
            "down_payment": round(down, 2),
            "loan_amount": round(loan, 2),
            "monthly_payment": round(monthly_payment, 2),
            "total_cost": round(monthly_payment * n_payments + down, 2),
            "current_liquid": round(current_savings, 2),
            "down_payment_gap": round(max(0, down - current_savings), 2),
        }

        self.create_scenario(Scenario(
            name="home_purchase",
            description=f"${target_price:,.0f} home at {interest_rate*100:.1f}% over {term_years}yr",
            assumptions={"price": target_price, "rate": interest_rate, "term": term_years},
            projections=result,
        ))
        return result

    # ------------------------------------------------------------------
    # Rocket Money sync
    # ------------------------------------------------------------------

    @property
    def rocket_money(self) -> RocketMoneyProvider:
        return self._rocket_money

    def sync_rocket_money(self) -> dict[str, Any]:
        """Pull accounts and transactions from Rocket Money (API or CSV).

        Merges synced data into the local ledger without duplicating
        transactions that already exist (matched by date + description + amount).
        """
        synced_accounts: list[SyncedAccount] = []
        synced_transactions: list[SyncedTransaction] = []

        if self._rm_connected:
            # API sync
            synced_accounts = self._rocket_money.fetch_accounts()
            now = datetime.now(timezone.utc)
            start = (now - timedelta(days=90)).isoformat()[:10]
            end = now.isoformat()[:10]
            synced_transactions = self._rocket_money.fetch_transactions(start, end)
        else:
            # CSV fallback: check data dir for Rocket Money CSVs
            csv_files = sorted(self._data_dir.glob("rocket_money*.csv"), reverse=True)
            if csv_files:
                result = self._rocket_money.sync_from_csv(csv_files[0])
                synced_accounts = self._rocket_money.fetch_accounts()
                now = datetime.now(timezone.utc)
                start = (now - timedelta(days=90)).isoformat()[:10]
                end = now.isoformat()[:10]
                synced_transactions = self._rocket_money.fetch_transactions(start, end)
                self.log("rocket_money_csv_sync", details=result)

        # Merge accounts (update balance if account exists, create if new)
        accounts_added = 0
        accounts_updated = 0
        for sa in synced_accounts:
            existing = self._accounts.get(sa.name)
            if existing:
                if sa.balance != 0.0:  # CSV doesn't have balances, skip zero
                    existing.balance = sa.balance
                    existing.last_synced = sa.last_updated
                    accounts_updated += 1
            else:
                try:
                    acct_type = AccountType(sa.account_type)
                except ValueError:
                    acct_type = AccountType.CHECKING
                self._accounts[sa.name] = Account(
                    name=sa.name,
                    account_type=acct_type,
                    balance=sa.balance,
                    institution=sa.institution,
                    last_synced=sa.last_updated,
                )
                accounts_added += 1

        # Merge transactions (deduplicate by date + description + amount)
        existing_keys = {
            (tx.date, tx.description, tx.amount) for tx in self._transactions
        }
        tx_added = 0
        for st in synced_transactions:
            key = (st.date, st.description, st.amount)
            if key not in existing_keys:
                try:
                    cat = TransactionCategory(st.category)
                except ValueError:
                    cat = TransactionCategory.OTHER
                self._transactions.append(Transaction(
                    date=st.date,
                    description=st.description,
                    amount=st.amount,
                    category=cat,
                    account=st.account,
                    metadata={"source": "rocket_money"},
                ))
                existing_keys.add(key)
                tx_added += 1

        self._last_sync = datetime.now(timezone.utc).isoformat()

        # Persist if anything changed
        if accounts_added or accounts_updated or tx_added:
            self.save_ledger()

        result = {
            "source": "api" if self._rm_connected else "csv",
            "accounts_added": accounts_added,
            "accounts_updated": accounts_updated,
            "transactions_added": tx_added,
            "total_accounts": len(self._accounts),
            "total_transactions": len(self._transactions),
            "synced_at": self._last_sync,
        }

        self.log("rocket_money_sync", details=result)
        return result

    def sync_from_csv(self, csv_path: str | Path) -> dict[str, Any]:
        """Directly import a Rocket Money CSV into the ledger.

        Use this when you have a CSV file (downloaded manually or via Gmail).
        """
        csv_result = self._rocket_money.sync_from_csv(csv_path)
        self.log("rocket_money_csv_loaded", details=csv_result)
        return self.sync_rocket_money()

    def sync_from_xlsx(self, xlsx_path: str | Path) -> dict[str, Any]:
        """Import a Rocket Money XLSX transaction export into the ledger.

        Expected columns (Rocket Money export format):
            Date, Original Date, Account Type, Account Name, Account Number,
            Institution Name, Name, Custom Name, Amount, Description,
            Category, Note, Ignored From, Tax Deductible, Transaction Tags

        Merges accounts and transactions, deduplicates, and persists.
        """
        from openpyxl import load_workbook
        from guardian_one.integrations.financial_sync import map_rocket_money_category

        wb = load_workbook(str(xlsx_path), read_only=True)
        ws = wb.active

        # Use iter_rows for efficient streaming (read_only mode)
        rows_iter = ws.iter_rows(values_only=True)
        headers = list(next(rows_iter))
        col = {h: i for i, h in enumerate(headers)}

        # Track accounts seen (to derive balances from transaction flow)
        account_info: dict[str, dict[str, Any]] = {}
        new_transactions: list[tuple[str, str, float, str, str]] = []  # date, desc, amt, cat, acct

        acct_type_map = {
            "cash": "checking", "credit card": "credit_card",
            "loan": "loan", "savings": "savings",
            "investment": "investment", "retirement": "retirement",
        }

        for row in rows_iter:
            date_val = row[col.get("Date", 0)]
            acct_type_raw = str(row[col.get("Account Type", 2)] or "cash").lower()
            acct_name = str(row[col.get("Account Name", 3)] or "Unknown")
            institution = str(row[col.get("Institution Name", 5)] or "Unknown")
            name = str(row[col.get("Name", 6)] or row[col.get("Custom Name", 7)] or "")
            amount = row[col.get("Amount", 8)]
            description = str(row[col.get("Description", 9)] or name)
            category_raw = str(row[col.get("Category", 10)] or "Other")

            if date_val is None or amount is None:
                continue

            # Format date
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)[:10]

            amount = float(amount)

            # Track account metadata
            if acct_name not in account_info:
                account_info[acct_name] = {
                    "institution": institution,
                    "type": acct_type_map.get(acct_type_raw, "checking"),
                }

            # Map category
            cfo_cat = map_rocket_money_category(category_raw)

            # Rocket Money: positive = outflow (spent), negative = inflow (income)
            # CFO convention: positive = inflow, negative = outflow
            cfo_amount = -amount

            new_transactions.append((date_str, description[:200], cfo_amount, cfo_cat, acct_name))

        wb.close()

        # Deduplicate against existing transactions
        existing_keys = {
            (tx.date, tx.description, tx.amount) for tx in self._transactions
        }

        tx_added = 0
        for date_str, desc, amt, cat, acct in new_transactions:
            key = (date_str, desc, amt)
            if key not in existing_keys:
                try:
                    tx_cat = TransactionCategory(cat)
                except ValueError:
                    tx_cat = TransactionCategory.OTHER
                self._transactions.append(Transaction(
                    date=date_str,
                    description=desc,
                    amount=amt,
                    category=tx_cat,
                    account=acct,
                    metadata={"source": "rocket_money_xlsx"},
                ))
                existing_keys.add(key)
                tx_added += 1

        # Update/add accounts
        accounts_added = 0
        accounts_updated = 0
        now_iso = datetime.now(timezone.utc).isoformat()
        for acct_name, info in account_info.items():
            # Skip Rocket Money goal buckets
            if info["institution"] == "Rocket Money":
                continue
            existing = self._accounts.get(acct_name)
            if existing:
                existing.last_synced = now_iso
                accounts_updated += 1
            else:
                try:
                    acct_type = AccountType(info["type"])
                except ValueError:
                    acct_type = AccountType.CHECKING
                self._accounts[acct_name] = Account(
                    name=acct_name,
                    account_type=acct_type,
                    balance=0.0,  # Will be computed from transactions
                    institution=info["institution"],
                    last_synced=now_iso,
                )
                accounts_added += 1

        self._last_sync = now_iso

        # Persist
        self.save_ledger()

        result = {
            "source": "xlsx",
            "file": str(xlsx_path),
            "transactions_in_file": len(new_transactions),
            "transactions_added": tx_added,
            "accounts_added": accounts_added,
            "accounts_updated": accounts_updated,
            "total_accounts": len(self._accounts),
            "total_transactions": len(self._transactions),
        }
        self.log("xlsx_import", details=result)
        return result

    # ------------------------------------------------------------------
    # Empower sync
    # ------------------------------------------------------------------

    @property
    def empower(self) -> EmpowerProvider:
        return self._empower

    def sync_empower(self) -> dict[str, Any]:
        """Pull retirement/investment accounts and transactions from Empower.

        Merges into the local ledger without duplicating existing data.
        """
        if not self._empower_connected:
            return {
                "source": "empower",
                "connected": False,
                "error": self._empower.last_error,
            }

        synced_accounts = self._empower.fetch_accounts()
        now = datetime.now(timezone.utc)
        start = (now - timedelta(days=90)).isoformat()[:10]
        end = now.isoformat()[:10]
        synced_transactions = self._empower.fetch_transactions(start, end)

        # Merge accounts
        accounts_added = 0
        accounts_updated = 0
        for sa in synced_accounts:
            existing = self._accounts.get(sa.name)
            if existing:
                existing.balance = sa.balance
                existing.last_synced = sa.last_updated
                accounts_updated += 1
            else:
                try:
                    acct_type = AccountType(sa.account_type)
                except ValueError:
                    acct_type = AccountType.INVESTMENT
                self._accounts[sa.name] = Account(
                    name=sa.name,
                    account_type=acct_type,
                    balance=sa.balance,
                    institution=sa.institution or "Empower",
                    last_synced=sa.last_updated,
                )
                accounts_added += 1

        # Merge transactions (deduplicate)
        existing_keys = {
            (tx.date, tx.description, tx.amount) for tx in self._transactions
        }
        tx_added = 0
        for st in synced_transactions:
            key = (st.date, st.description, st.amount)
            if key not in existing_keys:
                try:
                    cat = TransactionCategory(st.category)
                except ValueError:
                    cat = TransactionCategory.SAVINGS
                self._transactions.append(Transaction(
                    date=st.date,
                    description=st.description,
                    amount=st.amount,
                    category=cat,
                    account=st.account,
                    metadata={"source": "empower"},
                ))
                existing_keys.add(key)
                tx_added += 1

        if accounts_added or accounts_updated or tx_added:
            self.save_ledger()

        result = {
            "source": "empower",
            "connected": True,
            "accounts_added": accounts_added,
            "accounts_updated": accounts_updated,
            "transactions_added": tx_added,
            "synced_at": datetime.now(timezone.utc).isoformat(),
        }
        self.log("empower_sync", details=result)
        return result

    def empower_status(self) -> dict[str, Any]:
        """Current Empower connection status."""
        return {
            **self._empower.status(),
            "connected": self._empower_connected,
        }

    def rocket_money_status(self) -> dict[str, Any]:
        """Current Rocket Money connection and sync status."""
        rm_status = self._rocket_money.status()
        return {
            **rm_status,
            "connected": self._rm_connected,
            "last_ledger_sync": self._last_sync,
            "sync_mode": "api" if self._rm_connected else (
                "csv" if rm_status["csv_transactions"] > 0 else "offline"
            ),
        }

    # ------------------------------------------------------------------
    # Plaid sync (direct bank connections — READ-ONLY)
    # ------------------------------------------------------------------

    @property
    def plaid(self) -> PlaidProvider:
        return self._plaid

    def sync_plaid(self) -> dict[str, Any]:
        """Pull accounts and transactions from Plaid-connected banks.

        Plaid provides direct, read-only access to bank accounts
        (BofA, Wells Fargo, Capital One, etc.).  Merges into the
        local ledger without duplicating existing data.
        """
        if not self._plaid_connected or not self._plaid.connected_institutions:
            return {
                "source": "plaid",
                "connected": self._plaid_connected,
                "institutions": len(self._plaid.connected_institutions),
                "error": self._plaid.last_error if not self._plaid_connected else "No institutions linked",
            }

        synced_accounts = self._plaid.fetch_accounts()
        now = datetime.now(timezone.utc)
        start = (now - timedelta(days=90)).isoformat()[:10]
        end = now.isoformat()[:10]
        synced_transactions = self._plaid.fetch_transactions(start, end)

        # Merge accounts
        accounts_added = 0
        accounts_updated = 0
        for sa in synced_accounts:
            existing = self._accounts.get(sa.name)
            if existing:
                existing.balance = sa.balance
                existing.last_synced = sa.last_updated
                accounts_updated += 1
            else:
                try:
                    acct_type = AccountType(sa.account_type)
                except ValueError:
                    acct_type = AccountType.CHECKING
                self._accounts[sa.name] = Account(
                    name=sa.name,
                    account_type=acct_type,
                    balance=sa.balance,
                    institution=sa.institution,
                    last_synced=sa.last_updated,
                )
                accounts_added += 1

        # Merge transactions (deduplicate)
        existing_keys = {
            (tx.date, tx.description, tx.amount) for tx in self._transactions
        }
        tx_added = 0
        for st in synced_transactions:
            key = (st.date, st.description, st.amount)
            if key not in existing_keys:
                try:
                    cat = TransactionCategory(st.category)
                except ValueError:
                    cat = TransactionCategory.OTHER
                self._transactions.append(Transaction(
                    date=st.date,
                    description=st.description,
                    amount=st.amount,
                    category=cat,
                    account=st.account,
                    metadata={"source": "plaid"},
                ))
                existing_keys.add(key)
                tx_added += 1

        if accounts_added or accounts_updated or tx_added:
            self.save_ledger()

        result = {
            "source": "plaid",
            "connected": True,
            "institutions": len(self._plaid.connected_institutions),
            "accounts_added": accounts_added,
            "accounts_updated": accounts_updated,
            "transactions_added": tx_added,
            "synced_at": datetime.now(timezone.utc).isoformat(),
        }
        self.log("plaid_sync", details=result)
        return result

    def plaid_status(self) -> dict[str, Any]:
        """Current Plaid connection status."""
        return {
            **self._plaid.status(),
            "connected": self._plaid_connected,
        }

    # ------------------------------------------------------------------
    # Dashboard
    # ------------------------------------------------------------------

    def dashboard(self) -> dict[str, Any]:
        """Full financial snapshot."""
        return {
            "net_worth": self.net_worth(),
            "balances_by_type": self.balances_by_type(),
            "accounts": len(self._accounts),
            "upcoming_bills": [
                {"name": b.name, "amount": b.amount, "due": b.due_date}
                for b in self.upcoming_bills()
            ],
            "overdue_bills": [
                {"name": b.name, "amount": b.amount, "due": b.due_date}
                for b in self.overdue_bills()
            ],
            "spending_this_month": self.spending_summary(
                datetime.now(timezone.utc).isoformat()[:7]
            ),
            "income_this_month": self.income_summary(
                datetime.now(timezone.utc).isoformat()[:7]
            ),
            "budget_check": self.budget_check(),
            "budget_alerts": self.budget_alerts(),
            "net_worth_trend": self.net_worth_trend(months=6),
            "rocket_money": self.rocket_money_status(),
            "plaid": self.plaid_status(),
        }

    def generate_excel(
        self,
        output_path: str | Path | None = None,
        password: str | None = None,
        gmail_data: dict[str, Any] | None = None,
    ) -> Path:
        """Generate the Excel dashboard spreadsheet.

        Args:
            output_path: Where to save the file (default: data/dashboard.xlsx)
            password: Password-protect the workbook. When set, all sheets
                      are locked — visible but not editable without the password.
            gmail_data: Optional Gmail financial email data for the daily check.

        Returns the path to the generated .xlsx file.
        """
        from guardian_one.agents.cfo_dashboard import generate_dashboard
        path = output_path or (self._data_dir / "dashboard.xlsx")
        return generate_dashboard(self, path, password=password, gmail_data=gmail_data)

    def validation_report(self) -> dict[str, Any]:
        """Produce a detailed validation report for presentation.

        Shows every account with balance, categorised totals, net worth
        breakdown, bill status, and sync timestamps — suitable for
        review against the live Empower / Rocket Money dashboards.
        """
        now = datetime.now(timezone.utc).isoformat()

        # Account detail list
        account_details = []
        for a in self._accounts.values():
            account_details.append({
                "name": a.name,
                "type": a.account_type.value,
                "balance": a.balance,
                "institution": a.institution,
                "last_synced": a.last_synced,
            })

        # Totals by category
        type_totals = self.balances_by_type()
        assets = sum(v for k, v in type_totals.items() if k not in ("credit_card", "loan"))
        liabilities = sum(v for k, v in type_totals.items() if k in ("credit_card", "loan"))

        # Bills summary
        overdue = self.overdue_bills()
        upcoming = self.upcoming_bills(days=30)

        return {
            "report_generated": now,
            "net_worth": self.net_worth(),
            "total_assets": round(assets, 2),
            "total_liabilities": round(liabilities, 2),
            "balances_by_type": type_totals,
            "accounts": account_details,
            "account_count": len(self._accounts),
            "transaction_count": len(self._transactions),
            "bills": {
                "total": len(self._bills),
                "overdue": [{"name": b.name, "amount": b.amount, "due": b.due_date} for b in overdue],
                "upcoming_30d": [{"name": b.name, "amount": b.amount, "due": b.due_date, "auto_pay": b.auto_pay} for b in upcoming],
            },
            "tax_recommendations": self.tax_recommendations(),
            "rocket_money": self.rocket_money_status(),
            "empower": self.empower_status(),
            "plaid": self.plaid_status(),
            "ledger_path": str(self._ledger_path),
        }

    def sync_all(self) -> dict[str, Any]:
        """Run a full sync cycle — Plaid + Empower + Rocket Money — and return results."""
        results: dict[str, Any] = {
            "synced_at": datetime.now(timezone.utc).isoformat(),
        }

        # Plaid (direct bank connections — primary source)
        results["plaid"] = self.sync_plaid()

        # Empower (retirement accounts)
        results["empower"] = self.sync_empower()

        # Rocket Money (aggregator fallback)
        results["rocket_money"] = self.sync_rocket_money()

        # Updated totals
        results["net_worth"] = self.net_worth()
        results["account_count"] = len(self._accounts)
        results["transaction_count"] = len(self._transactions)

        # Record net worth snapshot for trend tracking
        self.record_net_worth()

        # Check budgets
        results["budget_alerts"] = self.budget_alerts()

        self.log("sync_all_complete", details=results)
        return results

    # ------------------------------------------------------------------
    # BaseAgent interface
    # ------------------------------------------------------------------

    def run(self) -> AgentReport:
        self._set_status(AgentStatus.RUNNING)
        alerts: list[str] = []
        recommendations: list[str] = []
        actions: list[str] = []

        # Rocket Money sync
        rm_status = self.rocket_money_status()
        if self._rm_connected:
            sync = self.sync_rocket_money()
            actions.append(
                f"Rocket Money API sync: {sync['accounts_added']} new accounts, "
                f"{sync['transactions_added']} new transactions."
            )
        elif rm_status["csv_transactions"] > 0 or list(self._data_dir.glob("rocket_money*.csv")):
            sync = self.sync_rocket_money()
            actions.append(
                f"Rocket Money CSV sync: {sync['accounts_added']} new accounts, "
                f"{sync['transactions_added']} new transactions."
            )
        else:
            if not self._rocket_money.has_credentials:
                recommendations.append(
                    "Set ROCKET_MONEY_API_KEY in .env for live account sync, "
                    "or export a CSV from Rocket Money and run: python main.py --csv path/to/export.csv"
                )
            else:
                alerts.append(
                    f"Rocket Money API connection failed: {rm_status['last_error']}"
                )

        # Empower sync
        if self._empower_connected:
            emp_sync = self.sync_empower()
            actions.append(
                f"Empower sync: {emp_sync['accounts_added']} new accounts, "
                f"{emp_sync['transactions_added']} new transactions."
            )
        elif self._empower.has_credentials:
            alerts.append(
                f"Empower connection failed: {self._empower.last_error}"
            )
        else:
            recommendations.append(
                "Set EMPOWER_API_KEY in .env to sync retirement accounts from Empower."
            )

        # Plaid sync (direct bank connections)
        if self._plaid_connected and self._plaid.connected_institutions:
            plaid_sync = self.sync_plaid()
            actions.append(
                f"Plaid sync: {plaid_sync.get('accounts_added', 0)} new accounts, "
                f"{plaid_sync.get('transactions_added', 0)} new transactions "
                f"from {plaid_sync.get('institutions', 0)} bank(s)."
            )
        elif not self._plaid.has_credentials:
            recommendations.append(
                "Set PLAID_CLIENT_ID and PLAID_SECRET in .env, then run "
                "'python main.py --connect' to link bank accounts (read-only)."
            )

        # Bill checks
        overdue = self.overdue_bills()
        if overdue:
            for b in overdue:
                alerts.append(f"OVERDUE: {b.name} — ${b.amount:.2f} was due {b.due_date}")

        upcoming = self.upcoming_bills()
        if upcoming:
            for b in upcoming:
                if not b.auto_pay:
                    alerts.append(f"Bill due soon: {b.name} — ${b.amount:.2f} on {b.due_date}")

        # Budget alerts
        budget_warnings = self.budget_alerts()
        alerts.extend(budget_warnings)

        # Record net worth for trend tracking
        self.record_net_worth()

        # Tax recs
        tax_recs = self.tax_recommendations()
        recommendations.extend(tax_recs)

        actions.append("Generated financial dashboard and tax recommendations.")
        self._set_status(AgentStatus.IDLE)

        return AgentReport(
            agent_name=self.name,
            status=AgentStatus.IDLE.value,
            summary=f"Net worth: ${self.net_worth():,.2f} | {len(self._accounts)} accounts | {len(overdue)} overdue bills.",
            actions_taken=actions,
            recommendations=recommendations,
            alerts=alerts,
            data=self.dashboard(),
        )

    def report(self) -> AgentReport:
        return AgentReport(
            agent_name=self.name,
            status=self.status.value,
            summary=f"Tracking {len(self._accounts)} accounts, {len(self._bills)} bills, {len(self._scenarios)} scenarios.",
            data=self.dashboard(),
        )
